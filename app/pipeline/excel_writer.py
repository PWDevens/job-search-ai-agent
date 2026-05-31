"""
app/pipeline/excel_writer.py — Pipeline Excel Workbook + User File Merge
=========================================================================

Two public functions:

  append_jobs_to_pipeline(jobs)
    → Appends AI-matched jobs to the internal PIPELINE_XLSX workbook.
      Creates the file if it doesn't exist; deduplicates by job ID.

  merge_new_jobs_to_user_file(user_file_path, new_jobs)
    → Reads the user's uploaded CSV/XLSX, normalises its headers,
      deduplicates against what already exists, and appends only
      genuinely new jobs back to the SAME file in the user's format.
      The user's original column order is preserved; any standard
      columns missing from their file are added on the right.

Schema for the internal pipeline workbook column order:
  Run Date | Rank | Match Score | Title | Company | Location | Salary |
  Date Posted | Date Found | Last Updated | Status | URL | Source | Job ID

Microsoft Fluent Design System colours:
  Header:    #0078D4 (MS blue), white bold text
  Alt rows:  #EFF6FC (light blue tint)
  Status cells: semantic colour per application_status value
"""
from __future__ import annotations
import logging
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import PIPELINE_XLSX, UPLOAD_FOLDER
from app.pipeline.normalizer import (
    normalize_headers,
    fuzzy_remap,
    normalise_status,
    STANDARD_COLS,
)

logger = logging.getLogger(__name__)


# ── Internal pipeline workbook columns ────────────────────────────────────────
PIPELINE_COLUMNS = [
    "Run Date",
    "Rank",
    "Match Score",
    "Title",
    "Company",
    "Location",
    "Salary",
    "Date Posted",
    "Date Found",
    "Last Updated",
    "Application Status",
    "URL",
    "Source",
    "Job ID",
]

# ── Microsoft Fluent colour palette ───────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="0078D4")   # MS primary blue
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Segoe UI", size=10)
BODY_FONT   = Font(name="Segoe UI", size=10)
ALT_FILL    = PatternFill("solid", fgColor="EFF6FC")   # light blue stripe

# Application status → Fluent semantic fill colour
STATUS_FILLS: Dict[str, str] = {
    "Have Not Applied":    "F3F2F1",   # neutral gray
    "Applied":             "DEECF9",   # light blue
    "Under Consideration": "F4ECF9",   # light purple
    "Interviewing":        "DFF6DD",   # light green
    "Offer Received":      "107C10",   # solid green (white font)
    "Rejected":            "FDE7E9",   # light red
    "Withdrawn":           "EDEBE9",   # mid gray
}
# Statuses that need white text because the fill is dark
STATUS_WHITE_FONT = {"Offer Received"}


# ─────────────────────────────────────────────────────────────────────────────
# Public API — internal pipeline workbook
# ─────────────────────────────────────────────────────────────────────────────

def append_jobs_to_pipeline(jobs: List[Dict[str, Any]]) -> int:
    """
    Append AI-matched jobs to the internal PIPELINE_XLSX workbook.

    - Creates the workbook with a styled header row if it doesn't exist.
    - Deduplicates by job ID (SHA-1 of document text) so re-running the
      pipeline never creates duplicate rows.
    - Preserves existing application_status values for jobs already present.
    - Applies Fluent status-colour fills to the Application Status column.

    Args:
        jobs: List of job dicts from find_top_jobs() / _format_results().

    Returns:
        Number of new rows written.
    """
    path = Path(PIPELINE_XLSX)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb, ws, existing_ids = _load_or_create_pipeline(path)
    run_date = datetime.now().strftime("%Y-%m-%d %H:%M")

    new_rows = 0
    for job in jobs:
        jid = job.get("id", "")
        if jid and jid in existing_ids:
            logger.debug("Skipping duplicate job_id=%s", jid)
            continue

        status = job.get("application_status", "Have Not Applied") or "Have Not Applied"
        status = normalise_status(status)

        row_data = [
            run_date,
            job.get("rank",       ""),
            _fmt_score(job.get("score", "")),
            job.get("title",      ""),
            job.get("company",    ""),
            job.get("location",   ""),
            job.get("salary",     ""),
            job.get("date_posted",""),
            job.get("date_found", ""),
            job.get("date_of_last_update", ""),
            status,
            job.get("url",        ""),
            job.get("source",     ""),
            jid,
        ]
        ws.append(row_data)
        new_rows += 1

        row_idx = ws.max_row
        _apply_body_row_style(ws, row_idx, new_rows, status)

        # Clickable hyperlink for URL
        url_val = job.get("url", "")
        if url_val:
            url_col_idx = PIPELINE_COLUMNS.index("URL") + 1
            cell = ws.cell(row=row_idx, column=url_col_idx)
            cell.hyperlink = url_val
            cell.font = Font(
                name="Segoe UI", size=10, color="0078D4", underline="single"
            )

    _auto_size_columns(ws)
    wb.save(path)
    logger.info("Pipeline workbook updated: +%d new rows → %s", new_rows, path)
    return new_rows


# ─────────────────────────────────────────────────────────────────────────────
# Public API — merge into user-uploaded file
# ─────────────────────────────────────────────────────────────────────────────

def merge_new_jobs_to_user_file(
    user_file_path: Path,
    new_jobs:       List[Dict[str, Any]],
) -> Tuple[int, Path]:
    """
    Append genuinely new AI-found jobs back to the user's uploaded file.

    Design:
    -------
    1. Read the user's file (CSV or XLSX) with full header normalisation.
    2. Build deduplication keys from existing rows:
         Primary:  normalised URL   (if non-empty)
         Fallback: normalised title + "|" + normalised company
    3. For each job returned by the pipeline:
         - Skip if its dedup key is already in the file.
         - Map the job dict to the file's column order.
         - Set application_status = "Have Not Applied" for new rows.
         - Set date_found = today, date_of_last_update = today.
         - Set source = "AI-Found" if not already set in the job dict.
    4. Append new rows to the DataFrame.
    5. Save back:
         - XLSX → openpyxl with Fluent styling for new rows.
         - CSV  → plain UTF-8 CSV (preserves original format).

    The user's original column order is preserved. Any standard columns
    absent from their file (e.g. they didn't have 'date_found') are
    added as empty columns to the right of their existing columns.

    Args:
        user_file_path: Path to the user's uploaded CSV or XLSX on disk.
        new_jobs:       List of job dicts from the CrewAI / matcher pipeline.

    Returns:
        Tuple of (count_appended, output_path).
        output_path is the same as user_file_path (written in place).
    """
    path = Path(user_file_path)
    if not path.exists():
        logger.warning("User file not found for merge: %s", path)
        return 0, path

    suffix = path.suffix.lower()

    # ── 1. Load and normalise user's file ────────────────────────────────────
    if suffix in {".xlsx", ".xls"}:
        df = pd.read_excel(path, dtype=str)
    else:
        df = pd.read_csv(path, dtype=str)

    original_display_cols = list(df.columns)   # human-readable originals (preserved)
    df = normalize_headers(df)
    df = fuzzy_remap(df)
    normalised_cols = list(df.columns)          # canonical names after normalisation

    # Fill NaN so string ops don't fail
    df = df.fillna("")

    # ── 2. Build dedup keys from existing rows ───────────────────────────────
    existing_keys = _build_existing_keys(df)
    logger.info(
        "Loaded %d existing rows from %s; %d unique dedup keys",
        len(df), path.name, len(existing_keys),
    )

    # ── 3. Ensure all standard columns exist (add blanks if missing) ─────────
    for col in STANDARD_COLS:
        if col not in df.columns:
            df[col] = ""

    # The user's original columns come first; any we added follow
    user_col_order = normalised_cols + [
        c for c in STANDARD_COLS if c not in normalised_cols
    ]
    # description is often not in a tracking sheet — keep it last so it
    # doesn't disrupt the user's layout
    if "description" in user_col_order and "description" not in normalised_cols:
        user_col_order.remove("description")
        user_col_order.append("description")

    today_str = date.today().isoformat()

    # ── 4. Build new rows, skip duplicates ───────────────────────────────────
    new_rows: List[Dict[str, str]] = []
    for job in new_jobs:
        job_key = _job_dedup_key(job)
        if job_key in existing_keys:
            logger.debug("Skipping existing job: %r", job_key)
            continue

        row: Dict[str, str] = {col: "" for col in user_col_order}

        # Map pipeline job dict → standard columns
        row["title"]               = str(job.get("title",       "") or "")
        row["company"]             = str(job.get("company",     "") or "")
        row["location"]            = str(job.get("location",    "") or "")
        row["description"]         = str(job.get("document",    "") or "")
        row["salary"]              = str(job.get("salary",      "") or "")
        row["url"]                 = str(job.get("url",         "") or "")
        row["date_posted"]         = str(job.get("date_posted", "") or "")
        row["date_found"]          = today_str
        row["date_of_last_update"] = today_str
        row["source"]              = str(job.get("source",      "") or "AI-Found")
        row["application_status"]  = "Have Not Applied"

        new_rows.append(row)

    if not new_rows:
        logger.info("No new jobs to append to %s (all duplicates).", path.name)
        return 0, path

    # ── 5. Append and save ───────────────────────────────────────────────────
    new_df = pd.DataFrame(new_rows)[user_col_order]
    merged_df = pd.concat([df[user_col_order], new_df], ignore_index=True)

    if suffix in {".xlsx", ".xls"}:
        # Re-build with Fluent styling; existing rows keep their original look,
        # new rows get status-colour fills.
        _write_styled_xlsx(
            merged_df,
            path,
            original_display_cols=original_display_cols,
            normalised_cols=normalised_cols,
            new_row_count=len(new_rows),
        )
    else:
        # CSV — restore original header names for existing columns
        col_rename = dict(zip(normalised_cols, original_display_cols))
        merged_df = merged_df.rename(columns=col_rename)
        merged_df.to_csv(path, index=False, encoding="utf-8-sig")

    logger.info(
        "Merged %d new job(s) into user file %s (%d total rows)",
        len(new_rows), path.name, len(merged_df),
    )
    return len(new_rows), path


# ─────────────────────────────────────────────────────────────────────────────
# Deduplication helpers
# ─────────────────────────────────────────────────────────────────────────────

def _build_existing_keys(df: pd.DataFrame) -> set:
    """
    Build a set of dedup keys from the rows already in the user's file.
    Key strategy (in priority order):
      1. Normalised URL  — most reliable unique identifier
      2. Normalised title + "|" + normalised company — fallback
    """
    keys = set()
    url_col    = "url"    if "url"    in df.columns else None
    title_col  = "title"  if "title"  in df.columns else None
    company_col= "company"if "company"in df.columns else None

    for _, row in df.iterrows():
        url = _norm_key(row.get(url_col, "")) if url_col else ""
        if url:
            keys.add(url)
        elif title_col and company_col:
            tc_key = (
                _norm_key(row.get(title_col, ""))
                + "|"
                + _norm_key(row.get(company_col, ""))
            )
            if tc_key and tc_key != "|":
                keys.add(tc_key)
    return keys


def _job_dedup_key(job: Dict[str, Any]) -> str:
    """Generate the same-style dedup key for a pipeline job dict."""
    url = _norm_key(job.get("url", "") or "")
    if url:
        return url
    title   = _norm_key(job.get("title",   "") or "")
    company = _norm_key(job.get("company", "") or "")
    return f"{title}|{company}"


def _norm_key(val: str) -> str:
    """Lowercase, strip whitespace — used for dedup comparisons only."""
    return str(val).strip().lower()


# ─────────────────────────────────────────────────────────────────────────────
# Pipeline workbook helpers
# ─────────────────────────────────────────────────────────────────────────────

def _load_or_create_pipeline(path: Path):
    """Load an existing pipeline workbook or create a fresh one."""
    existing_ids: set = set()

    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # Find the Job ID column index
        job_id_col = None
        for cell in ws[1]:
            if cell.value == "Job ID":
                job_id_col = cell.column
                break
        if job_id_col:
            for row in ws.iter_rows(min_row=2, min_col=job_id_col,
                                     max_col=job_id_col, values_only=True):
                if row[0]:
                    existing_ids.add(row[0])
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Job Pipeline"
        ws.append(PIPELINE_COLUMNS)

        for col_idx, cell in enumerate(ws[1], 1):
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 20

    return wb, ws, existing_ids


def _apply_body_row_style(ws, row_idx: int, row_num: int, status: str) -> None:
    """Apply alternating row fill and status-colour to a data row."""
    status_hex = STATUS_FILLS.get(status, "FFFFFF")
    status_col_idx = PIPELINE_COLUMNS.index("Application Status") + 1

    for col_idx, cell in enumerate(ws[row_idx], 1):
        cell.font = BODY_FONT
        cell.alignment = Alignment(vertical="center")
        # Alternating stripe on non-status columns
        if col_idx != status_col_idx:
            if row_num % 2 == 0:
                cell.fill = ALT_FILL
        else:
            # Status column: semantic colour
            cell.fill = PatternFill("solid", fgColor=status_hex)
            if status in STATUS_WHITE_FONT:
                cell.font = Font(
                    name="Segoe UI", size=10, color="FFFFFF", bold=True
                )


def _fmt_score(score) -> str:
    """Format cosine similarity score as percentage string."""
    try:
        return f"{float(score):.1%}"
    except (TypeError, ValueError):
        return str(score)


def _auto_size_columns(ws, max_width: int = 50) -> None:
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        try:
            width = max(
                len(str(cell.value)) for cell in col_cells if cell.value is not None
            )
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(width + 3, max_width)
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────────────
# Styled XLSX write for merged user file
# ─────────────────────────────────────────────────────────────────────────────

def _write_styled_xlsx(
    merged_df:            pd.DataFrame,
    path:                 Path,
    original_display_cols: List[str],
    normalised_cols:      List[str],
    new_row_count:        int,
) -> None:
    """
    Write merged_df to path as a styled XLSX workbook.

    - Existing rows (everything except the last new_row_count rows) keep
      plain Segoe UI styling.
    - New rows (the appended leads) get status-colour fills in the
      Application Status column and alternating row stripe.
    - Column headers are restored to the user's original display names
      for columns the user already had; new columns keep canonical names.
    """
    # Build header display names: original names first, then canonical for new cols
    col_display: List[str] = []
    norm_to_orig = dict(zip(normalised_cols, original_display_cols))
    for col in merged_df.columns:
        col_display.append(norm_to_orig.get(col, col))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Tracker"

    # Header row
    ws.append(col_display)
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    # Determine status column index (0-based in df, 1-based in xlsx)
    status_col_xlsx = None
    for i, col in enumerate(merged_df.columns):
        if col == "application_status":
            status_col_xlsx = i + 1
            break

    total_rows = len(merged_df)
    first_new_row_idx = total_rows - new_row_count   # 0-based index in df

    for df_row_idx, (_, row) in enumerate(merged_df.iterrows()):
        row_values = [str(v) if v not in ("", None) else "" for v in row]
        ws.append(row_values)

        xlsx_row_idx = df_row_idx + 2  # +1 for header, +1 for 1-based
        is_new = df_row_idx >= first_new_row_idx
        status_val = str(row.get("application_status", "")).strip()

        for col_idx, cell in enumerate(ws[xlsx_row_idx], 1):
            cell.font      = BODY_FONT
            cell.alignment = Alignment(vertical="center")

            if status_col_xlsx and col_idx == status_col_xlsx:
                status_hex = STATUS_FILLS.get(status_val, "FFFFFF")
                cell.fill  = PatternFill("solid", fgColor=status_hex)
                if status_val in STATUS_WHITE_FONT:
                    cell.font = Font(
                        name="Segoe UI", size=10, color="FFFFFF", bold=True
                    )
            elif is_new and df_row_idx % 2 == 0:
                cell.fill = ALT_FILL

        # Hyperlink for URL column
        url_col_idx = None
        for i, col in enumerate(merged_df.columns):
            if col == "url":
                url_col_idx = i + 1
                break
        if url_col_idx:
            url_val = str(row.get("url", "")).strip()
            if url_val.startswith("http"):
                url_cell = ws.cell(row=xlsx_row_idx, column=url_col_idx)
                url_cell.hyperlink = url_val
                url_cell.font = Font(
                    name="Segoe UI", size=10, color="0078D4", underline="single"
                )

    _auto_size_columns(ws)
    wb.save(path)
